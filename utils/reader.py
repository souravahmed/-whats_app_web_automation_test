from openpyxl import load_workbook

from utils.file_path import FilePathUtil
import os


class ExcelUtil:
    
    @classmethod
    def open(self):
        try:
            self.file_path = os.path.join(FilePathUtil.get_excel_file_path(), "mobile-numbers.xlsx")
            self.wb = load_workbook(self.file_path)
            self.ws = self.wb.active
        except FileNotFoundError:
            print("could not found the mobile-numbers.xlsx ")
            
            
    @classmethod
    def get_mobile_number(self):
        return self.ws.cell(2, 1).value
    
    @classmethod
    def write_message(self, row, col, message):
        self.ws.cell(row, col).value = message
        self.wb.save(self.file_path)
        
        